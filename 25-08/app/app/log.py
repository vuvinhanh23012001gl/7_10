import logging
import os

class Log:
    from folder_create import Create
    obj_folder = Create()
    characters_check = "log_"
    def __init__(self,obj_config_software=None,name="app"):
        self.name = name
        self.console_enabled = False
        self.file_enabled = False
        self.logger = logging.getLogger(name)
        self.logger.setLevel(logging.DEBUG)  # Cho phép tất cả mức log
        self.logger.handlers.clear()  # Xóa handler cũ tránh bị log trùng
        self.formatter = logging.Formatter(
            "%(asctime)s [%(levelname)s]:%(message)s"
        )
        self.obj_config_software = obj_config_software
        if self.obj_config_software :
            self.log_file = self.create_file_log_software()
            if self.log_file:
                self.enable_file()
            else:
                self.disable_file()
            self.delete_file_old_log_software()
        self.enable_console()
        
    def get_time_software(self):
        """Lấy thời gian cho phép log phan mem được lưu nếu được bật"""
        return self.obj_config_software.GetTimeSaveLogSoftware()

    def get_open_log_software(self):
        """lấy quyền lưu log"""
        return self.obj_config_software.get_log_software()

    def get_path_save_software(self):
        """Tra ve duong dan luu Fodel log"""
        return self.obj_config_software.get_path_log_software()
    

    def get_list_file_in_folder_log_sofware(self)->list:
        """Hàm này trả về danh sách file hiện có trong folder software"""
        return Log.obj_folder.get_list_file_in_folder(self.get_path_save_software())
    
    def get_list_find_old_sofware(self,days_threshold):
        """Trả về đường danh sách tên file có days_threshold không thỏa mãn để xóa """
        list_file = self.get_list_file_in_folder_log_sofware()
        print("Danh sach file excell hiện có trong thư mục là:",list_file)
        if  list_file:
            arr_old_file  = Log.obj_folder.get_old_files_by_threshold(Log.characters_check,list_file,days_threshold,"txt")
            print(f"Danh sách file cũ hơn {days_threshold} ngày để xóa",arr_old_file)
            return arr_old_file
        else :
            print("Danh sách trong folder excell rỗng")
            return None
    def delete_file_old_log_software(self):
        arr_file_old = self.get_list_find_old_sofware(self.get_time_software())
        print("arr_file_old",arr_file_old)
        if arr_file_old:
            print("---Xóa File quá hạn \n Bắt đầu xóa --")
            for file_delete in arr_file_old:
                path_file_delete = Log.obj_folder.find_file(self.get_path_save_software(),file_delete)
                print(path_file_delete)
                if path_file_delete:
                    Log.obj_folder.delete_file(path_file_delete)
            print("--Xóa thành công file--")
    def create_file_log_software(self):
       """Trả về đường dẫn của file log sản phẩn nếu folder không có trả về None"""
       open_log_software = self.get_open_log_software()
       if open_log_software:
            path_log_folder_software = self.obj_config_software.get_path_log_software()
            print("Đường Link log tồn tại .")
            print(path_log_folder_software)
            return Log.obj_folder.create_file_text_log(path_log_folder_software,"txt")
       return None
   
    def log_and_print(self, msg, value=None, level="info"):
        # Ghép message nếu có value
        full_msg = f"{msg}: {value}" if value is not None else msg
        if level == "debug":
            self.logger.debug(full_msg)
        elif level == "warning":
            self.logger.warning(full_msg)
        elif level == "error":
            self.logger.error(full_msg)
        elif level == "critical":
            self.logger.critical(full_msg)
        else:
            self.logger.info(full_msg)
            
    def enable_console(self):
            if not self.console_enabled:
                ch = logging.StreamHandler()
                ch.setLevel(logging.DEBUG)
                ch.setFormatter(self.formatter)
                self.logger.addHandler(ch)
                self.console_enabled = True

    def disable_console(self):
            for h in list(self.logger.handlers):
                if isinstance(h, logging.StreamHandler):
                    self.logger.removeHandler(h)
            self.console_enabled = False

    def enable_file(self):
            print("Bật Log File")
            if not self.file_enabled:
                os.makedirs(os.path.dirname(self.log_file) or ".", exist_ok=True)
                fh = logging.FileHandler(self.log_file, encoding="utf-8")
                fh.setLevel(logging.DEBUG)
                fh.setFormatter(self.formatter)
                self.logger.addHandler(fh)
                self.file_enabled = True

    def disable_file(self):
            print("Tắt Log File")
            print("TATA LOGGGGGGGGGGGGGG  fILE")
            for h in list(self.logger.handlers):
                if isinstance(h, logging.FileHandler):
                    self.logger.removeHandler(h)
            self.file_enabled = False

    # ===============================
    # Các hàm log tiện dụng
    # ===============================
    def debug(self, msg):
        self.logger.debug(msg)
        

    def info(self, msg):
        self.logger.info(msg)

    def warning(self, msg):
        self.logger.warning(msg)

    def error(self, msg):
        self.logger.error(msg)

    def critical(self, msg):
        self.logger.critical(msg)



# from config_software import OilDetectionSystem
# obj_config_software = OilDetectionSystem()
# obj_log_data = Log(obj_config_software)
# obj_log_data.info("wewewe232323232we")
# obj_log_data.info("wewewe232323232we")
# obj_log_data.info("wewewe232323232we")
# obj_log_data.info("wewewe232323232we")
# path_file  = obj_log_data.create_file_log_software()
# print(path_file)
# obj_log_data.delete_file_old_log_software()
#lAY THOI GIAN LUU LOG SOFTWARE
# print(obj_log_data.get_time_software())
# print(obj_log_data.get_open_log_software())
# print(obj_log_data.get_path_save_software())
# print(obj_log_data.get_list_file_in_folder_log_sofware())
# print(obj_log_data.get_list_find_old_sofware(1))



from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

class log_excell:
    '''Lớp này mỗi khi gọi sẽ tạo ra 1 file logexcell nếu được bật log excell ở trong config nếu tắt thì sẽ không tự động tạo ra 1 file'''
    from folder_create import Create
    obj_folder = Create()
    characters_check = "log_"  
    def __init__(self,obj_config_software):

        self.wb = None
        self.ws = None

        self.obj_config_software = obj_config_software
        

        self.path_file_save_log_excell = self.create_file_excell()  # self.path_file_save_log_excell  se cho co the bang none neu khong duoc phep tao file
        if self.path_file_save_log_excell:
            self.delete_file_old()
            print("Cho phép mở tạo và tạo file excell mới vì đang bật log excell tại đường dẫn",self.path_file_save_log_excell)
            self.write_file_excel(["Thời gian","Mã sản phẩm","Tên sản phẩm","Tên người thao tác","Mã lỗi","Ghi chú"])
        
    def get_path_file_save_log_excell(self):
        """Tra ve path File luu log hien tai"""
        return self.path_file_save_log_excell
    def get_time(self):
        """Lấy thời gian cho phép log được lưu nếu được bật"""
        return self.obj_config_software.GetTimeSaveLogExcell()

    def get_open_log_excell(self):
        """lấy quyền lưu log"""
        return  self.obj_config_software.get_log_product()

    def get_path_save_log_excell(self):
        return self.obj_config_software.get_path_log_product()
    

    def get_list_file_in_folder_log_excell(self)->list:
        """Hàm này trả về danh sách file hiện có trong folder excell"""
        return log_excell.obj_folder.get_list_file_in_folder(self.get_path_save_log_excell())
    
    def create_file_excell(self):
        """name_file la duong dan toi file"""
        if self.get_open_log_excell():
            file_path = log_excell.obj_folder.create_file_log(self.get_path_save_log_excell())
            return file_path
        else:
            print("Không tạo File Log sản phẩm vì không bật log")
            return None
    
    def get_list_find_old(self,days_threshold):
        """Trả về đường danh sách tên file có days_threshold không thỏa mãn để xóa """
        list_file = self.get_list_file_in_folder_log_excell()
        if  list_file:
            arr_old_file = log_excell.obj_folder.get_old_files_by_threshold(log_excell.characters_check,list_file,days_threshold)
            print(f"Danh sách file cũ hơn {days_threshold} ngày để xóa",arr_old_file)
            return arr_old_file
        else :
            print("Danh sách trong folder excell rỗng")
            return None
    def delete_file_old(self):
        arr_file_old = self.get_list_find_old(self.get_time())
        if arr_file_old:
            print("---Xóa File quá hạn \n Bắt đầu xóa --")
            for file_delete in arr_file_old:
                path_file_delete = log_excell.obj_folder.find_file(self.get_path_save_log_excell(),file_delete)
                print(path_file_delete)
                if path_file_delete:
                    log_excell.obj_folder.delete_file(path_file_delete)
            print("--Xóa thành công file--")


    def write_file_excel(self, row: list):
        """
        Ghi 1 dòng dữ liệu vào file Excel hiện tại.
        Nếu file chưa tồn tại -> tạo mới.
        :param row: list chứa dữ liệu tương ứng 1 dòng
        """
        if not self.get_open_log_excell():
            print("⚠️ Chức năng log chưa được bật, không lưu Excel")
            return None

        file_path = self.get_path_file_save_log_excell()

        # 🔹 Nếu file chưa tồn tại -> tạo mới
        if not os.path.exists(file_path):
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "Log Data"
            self.wb.save(file_path)
        else:
            self.wb = load_workbook(file_path)
            self.ws = self.wb.active

        # 🔹 Ghi dữ liệu vào dòng mới
        self.ws.append(row)

        # 🔹 Tự động căn chỉnh độ rộng cột cho đẹp
        for col in self.ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 4
            self.ws.column_dimensions[col_letter].width = adjusted_width

        # 🔹 Lưu file lại
        self.wb.save(file_path)
        # print(f"✅ Đã lưu dòng dữ liệu vào: {file_path}")

        return file_path
    
# from config_software import OilDetectionSystem
# obj_config_software = OilDetectionSystem()    
# test_obj_log_excell = log_excell(obj_config_software)

# print("Đường dẫn File Excell có nếu Bật log Excell",test_obj_log_excell.get_path_file_save_log_excell())
# test_obj_log_excell.get_time()
# print("Cho phéo tạo file không ?",test_obj_log_excell.get_open_log_excell())
# print("Dường dẫn lưu File excelc",test_obj_log_excell.get_path_save_log_excell())
# test_obj_log_excell.write_file_excel([1,3,4,5,23])
# test_obj_log_excell.delete_file_old()
# test_obj_log_excell.create_file_excell()
# test_obj_log_excell.get_list_find_old(1)



class log_img:
    #Kiểm thử hảm Log img Ok không cần kiểm tra lại
    from folder_create import Create
    obj_folder = Create()
    characters_check  = "img_"
    extension ="jpg"  
    def __init__(self,obj_config_software):
        self.obj_config_software = obj_config_software
        print("-------Tiến hành xóa File log ảnh-----")
        self.delete_file_old_log_img()  #xoa truoc moi khi mo phan mem
        print("-----------Hoàn thành xóa-----------")
    def get_path_foldef_log_img(self):
        """Tra ve path File luu log hien tai"""
        return self.obj_config_software.get_path_log_img_oil()
    def get_time_log_img(self):
        """Lấy thời gian cho phép log được lưu nếu được bật"""
        return self.obj_config_software.GetTimeSaveLogImg()
    def get_open_log_img(self):
        """lấy quyền lưu log"""
        return self.obj_config_software.get_log_img_oil()
    def create_file_log_img(self,img):
        """Hàm này tạo lưu ảnh img khi yêu cầu bật ảnh được bật"""
        if self.get_open_log_img():
            path_foler_img = self.get_path_foldef_log_img()
            log_img.obj_folder.create_file_log_img(img,path_foler_img,extension= log_img.extension)
    def get_list_find_old_img(self):
            """Trả về danh sách sau khoảng thời gian time trong cấu hình information"""
            time_set  = self.get_time_log_img()
            list_file =  self.get_list_file_in_folder_img()
            return log_img.obj_folder.get_old_files_by_threshold_img(log_img.characters_check,list_file,time_set)
    def delete_file_old_log_img(self):
        """Tự lấy danh sách ảnh cũ trong đường link ảnh và tự động xóa ảnh sau thời gian quá hạn"""
        arr_file_old_img = self.get_list_find_old_img()
        if arr_file_old_img:
            for file_delete in arr_file_old_img:
                path_file_delete = Log.obj_folder.find_file(self.get_path_foldef_log_img(),file_delete)
                if path_file_delete:
                    log_img.obj_folder.delete_file(path_file_delete)
            print("-------Xóa thành công file-------")
        print("Hiện tại không tìm thấy File quá hạn")
    def get_list_file_in_folder_img(self):
        return log_img.obj_folder.get_list_file_in_folder(self.get_path_foldef_log_img())

        
        


# from config_software import OilDetectionSystem
# obj_config_software = OilDetectionSystem()
# obj_log_img = log_img(obj_config_software)

# print("Thời gian lưu log hình ảnh hiện tại",obj_log_img.get_time_log_img(),"ngày")
# print("Đường dẫn lưu ảnh:",obj_log_img.get_path_foldef_log_img())
# print("Danh sách ảnh cũ quá hạn",obj_log_img.get_list_find_old_img())
# print("Danh sách ảnh cũ quá hạn",obj_log_img.get_list_file_in_folder_img())

# import numpy as np
# import os
# height, width, channels = 480, 640, 3
# blank_image = np.zeros((height, width, channels), dtype=np.uint8)  
# obj_log_img.create_file_log_img(blank_image)
# obj_log_img.delete_file_old_log_img()




